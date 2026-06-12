"""엑셀 저장 에이전트 — 차트·집계 조회 데이터를 .xlsx로보내기."""
from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXPORT_DIR = Path(__file__).resolve().parent / "exports"

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*]')


def _safe_sheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", (name or "데이터").strip()) or "데이터"
    return cleaned[:31]


def safe_excel_filename(name: str) -> str:
    base = _INVALID_FILENAME_CHARS.sub("_", (name or "culture_export").strip()) or "culture_export"
    if not base.lower().endswith(".xlsx"):
        base += ".xlsx"
    return base[:120]


def ascii_export_filename(prefix: str = "culture_export") -> str:
    """브라우저·Windows 다운로드용 ASCII 파일명."""
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", (prefix or "culture_export").strip()) or "culture_export"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{safe}_{ts}.xlsx"


def save_excel_to_disk(content: bytes, filename: str) -> Path:
    """로컬 PC culture/exports 폴더에 xlsx 저장."""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORT_DIR / safe_excel_filename(filename)
    path.write_bytes(content)
    return path.resolve()


def export_has_data(export: dict[str, Any]) -> bool:
    if export.get("rows"):
        return True
    for sheet in export.get("sheets") or []:
        if sheet.get("rows"):
            return True
    return False


def _write_sheet(ws, sheet_spec: dict[str, Any]) -> None:
    columns: list[str] = list(sheet_spec.get("columns") or [])
    rows: list[dict[str, Any]] = list(sheet_spec.get("rows") or [])
    if not columns and rows:
        columns = list(rows[0].keys())
    if not columns:
        raise ValueError("엑셀로 저장할 컬럼이 없습니다.")

    header_fill = PatternFill("solid", fgColor="004B8D")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title = (sheet_spec.get("title") or "").strip()
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        if len(columns) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        start_row = 2

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    for col_idx in range(1, len(columns) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(columns[col_idx - 1]))
        for row in rows:
            max_len = max(max_len, len(str(row.get(columns[col_idx - 1], ""))))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)


def build_excel_bytes(export: dict[str, Any]) -> bytes:
    """export payload → xlsx bytes (단일 시트 또는 sheets 배열)."""
    sheets = list(export.get("sheets") or [])
    if not sheets:
        sheets = [
            {
                "title": export.get("title"),
                "sheet_name": export.get("sheet_name") or "데이터",
                "columns": export.get("columns") or [],
                "rows": export.get("rows") or [],
            }
        ]

    wb = Workbook()
    wb.remove(wb.active)
    for idx, sheet_spec in enumerate(sheets):
        if not sheet_spec.get("rows"):
            continue
        name = _safe_sheet_name(str(sheet_spec.get("sheet_name") or f"Sheet{idx + 1}"))
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, sheet_spec)

    if not wb.sheetnames:
        raise ValueError("엑셀로 저장할 데이터가 없습니다.")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
